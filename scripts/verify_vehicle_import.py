"""本地验证：车辆批量导入模板 → 落库 → 触发 jt808 upsert_now。

用法（在 backend 目录）：
  python scripts/verify_vehicle_import.py
"""
from __future__ import annotations

import asyncio
import sys
import tempfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


async def main() -> int:
    from app.models import Base, Fleet, OrgCompany, Vehicle, VehicleDevice
    from app.routers import api_vehicle
    from app import jt808_vehicle

    tmp = Path(tempfile.mkdtemp(prefix="cesg_vehicle_import_"))
    db_path = tmp / "test.db"
    engine = create_async_engine(f"sqlite+aiosqlite:///{db_path.as_posix()}", echo=False)
    Session = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    sync_calls: list[tuple[int, str | None]] = []

    async def fake_upsert(vehicle_id: int, old_device_no: str | None = None):
        sync_calls.append((int(vehicle_id), old_device_no))
        return True

    orig_upsert = jt808_vehicle.upsert_now
    jt808_vehicle.upsert_now = fake_upsert  # type: ignore[assignment]

    try:
        # 1) 模板表头
        tmpl_resp = await api_vehicle.download_vehicle_import_template()
        body = b"".join([chunk async for chunk in tmpl_resp.body_iterator])
        wb = load_workbook(BytesIO(body))
        headers = [str(c.value or "").strip() for c in next(wb.active.iter_rows(min_row=1, max_row=1))]
        assert headers == api_vehicle._VEHICLE_IMPORT_HEADERS, headers
        assert "填写说明" in wb.sheetnames
        print("[ok] template headers")

        # 2) 种子公司/车队
        async with Session() as db:
            co = OrgCompany(name="验证导入公司", org_code="VIMP01", jt808_group_id=90001)
            db.add(co)
            await db.flush()
            fl = Fleet(company_id=co.id, name="一车队")
            db.add(fl)
            await db.commit()
            company_id = co.id

        # 3) 构造合法导入文件
        plate = "测Z90001"
        device_no = "900000000001"
        wb2 = Workbook()
        ws = wb2.active
        ws.title = "车辆信息导入"
        ws.append(list(api_vehicle._VEHICLE_IMPORT_HEADERS))
        ws.append(
            [
                plate,
                "黄牌",
                "LSVTESTVIN0000001",
                "ENG-IMP-1",
                "PMC-IMP-1",
                "FRAME-IMP-1",
                "验证导入公司",
                "一车队",
                device_no,
                "JT808",
                "13800000001",
                "燃油车",
                "测试品牌",
                "测试型号",
                "正常",
                "2026-01-01",
                "2027-01-01",
                "2026-01-02",
                "4",
            ]
        )
        bio = BytesIO()
        wb2.save(bio)
        payload = bio.getvalue()

        class _Upload:
            filename = "vehicle_import_test.xlsx"

            async def read(self):
                return payload

        async with Session() as db:
            result = await api_vehicle.import_vehicle_from_carinfos(_Upload(), db=db)

        assert result["ok"] is True
        assert result["imported"] == 1, result
        assert result["updated"] == 0, result
        assert result["skipped"] == 0, result
        assert result["jt808_sync_ok"] == 1, result
        assert len(sync_calls) == 1 and sync_calls[0][0] > 0
        print("[ok] import + 808 upsert called:", sync_calls[0])

        async with Session() as db:
            v = await db.scalar(select(Vehicle).where(Vehicle.plate_no == plate).limit(1))
            assert v is not None
            assert v.company_id == company_id
            assert v.fleet_id is not None
            assert v.vin == "LSVTESTVIN0000001"
            assert v.engine_no == "ENG-IMP-1"
            assert v.product_model_code == "PMC-IMP-1"
            assert v.frame_no == "FRAME-IMP-1"
            assert v.vehicle_category == "fuel"
            assert v.brand == "测试品牌"
            d = await db.scalar(
                select(VehicleDevice).where(
                    VehicleDevice.vehicle_id == v.id, VehicleDevice.is_main.is_(True)
                ).limit(1)
            )
            assert d is not None
            assert d.device_no == device_no
            assert d.sim_no == "13800000001"
        print("[ok] db fields")

        # 4) 缺必填应跳过且不半写入
        sync_calls.clear()
        wb3 = Workbook()
        ws3 = wb3.active
        ws3.append(list(api_vehicle._VEHICLE_IMPORT_HEADERS))
        ws3.append(["测Z90002", "黄牌", "", "", "", "", "验证导入公司", "", "900000000002"] + [""] * 10)
        bio3 = BytesIO()
        wb3.save(bio3)

        class _UploadBad:
            filename = "bad.xlsx"

            async def read(self):
                return bio3.getvalue()

        async with Session() as db:
            bad = await api_vehicle.import_vehicle_from_carinfos(_UploadBad(), db=db)
        assert bad["imported"] == 0
        assert bad["skipped"] >= 1
        assert bad.get("annotated_file_base64")
        async with Session() as db:
            assert await db.scalar(select(Vehicle).where(Vehicle.plate_no == "测Z90002").limit(1)) is None
        print("[ok] skip invalid row without writing")

        print("ALL PASSED")
        return 0
    finally:
        jt808_vehicle.upsert_now = orig_upsert  # type: ignore[assignment]
        await engine.dispose()


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
