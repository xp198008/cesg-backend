"""按 docs/CAR.XLSX 对齐 CESG 车辆，再同步 808，最后删除表外多余车辆。

用法（在 backend 目录，现网用 venv python）：
  python scripts/reconcile_vehicles_from_car_xlsx.py --xlsx /path/CAR.XLSX --dry-run
  python scripts/reconcile_vehicles_from_car_xlsx.py --xlsx /path/CAR.XLSX --apply
"""
from __future__ import annotations

import argparse
import asyncio
import json
import logging
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

logger = logging.getLogger("reconcile_car_xlsx")

@dataclass(frozen=True)
class CompanyRef:
    id: int
    name: str
    org_code: str


COL_ALIASES = {
    "plate": ("车牌号",),
    "vin": ("车架号", "VIN", "车辆识别代码VIN", "车辆识别代码"),
    "company": ("所属公司",),
    "parent_company": ("上级公司",),
    "fleet": ("车队",),
    "install_date": ("安装日期",),
    "category": ("能源类型", "车辆类别", "燃料类型"),
    "brand": ("车辆品牌", "品牌"),
    "model": ("车辆型号", "型号"),
    "channel": ("通道数目", "通道数"),
    "device": ("设备编号", "设备号", "设备1编号"),
    "sim": ("设备SIM", "SIM卡号", "设备1SIM", "设备SIM卡号"),
}


def _norm(s: Any) -> str:
    if s is None:
        return ""
    if isinstance(s, datetime):
        return s.strftime("%Y-%m-%d")
    if isinstance(s, date):
        return s.isoformat()
    return str(s).replace("\u3000", " ").strip()


def _norm_plate(s: Any) -> str:
    return _norm(s).replace(" ", "").upper()


def _to_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _norm(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_int(v: Any) -> int | None:
    s = _norm(v)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _norm_category(raw: Any) -> str | None:
    t = _norm(raw)
    if not t:
        return None
    low = t.lower()
    if low in ("fuel", "燃油", "燃油车", "油", "柴油", "汽油"):
        return "fuel"
    if low in ("new", "新能源", "新能源车", "ev", "电动", "电动车", "电"):
        return "new"
    return t


def _header_map(headers: list[str]) -> dict[str, int]:
    found: dict[str, int] = {}
    for key, aliases in COL_ALIASES.items():
        for name in aliases:
            if name in headers:
                found[key] = headers.index(name)
                break
    missing = [k for k in ("plate", "company", "device") if k not in found]
    if missing:
        raise SystemExit(f"Excel 缺少必要列：{missing}；实际表头={headers}")
    return found


def _cell(row: tuple, idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def load_excel_rows(xlsx: Path) -> list[dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx), data_only=True, read_only=True)
    ws = wb.active
    it = ws.iter_rows(min_row=1, values_only=True)
    header_row = next(it, ())
    headers = [_norm(v) for v in header_row]
    col = _header_map(headers)
    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for ridx, row in enumerate(it, start=2):
        plate = _norm_plate(_cell(row, col["plate"]))
        if not plate:
            continue
        if plate in seen:
            raise SystemExit(f"Excel 第 {ridx} 行车牌重复：{plate}")
        seen.add(plate)
        out.append(
            {
                "row": ridx,
                "plate": plate,
                "vin": _norm(_cell(row, col.get("vin"))),
                "company": _norm(_cell(row, col["company"])),
                "parent_company": _norm(_cell(row, col.get("parent_company"))),
                "fleet": _norm(_cell(row, col.get("fleet"))),
                "install_date": _to_date(_cell(row, col.get("install_date"))),
                "category": _norm_category(_cell(row, col.get("category"))),
                "brand": _norm(_cell(row, col.get("brand"))),
                "model": _norm(_cell(row, col.get("model"))),
                "channel": _to_int(_cell(row, col.get("channel"))),
                "device": _norm(_cell(row, col["device"])),
                "sim": _norm(_cell(row, col.get("sim"))),
            }
        )
    wb.close()
    return out


def _set_if(obj: Any, field: str, value: Any, changes: list[str]) -> None:
    if value is None:
        return
    if isinstance(value, str) and not value:
        return
    old = getattr(obj, field)
    if old != value:
        setattr(obj, field, value)
        changes.append(f"{field}:{old!r}->{value!r}")


async def _delete_cesg_vehicle(db: AsyncSession, vehicle_id: int) -> tuple[str, str | None]:
    from app.models import Vehicle, VehicleAllocRuleVehicle, VehicleDevice, VehicleLocation
    from app import jt808_vehicle

    row = await db.scalar(select(Vehicle).where(Vehicle.id == vehicle_id).limit(1))
    if row is None:
        return "", None
    plate_no = row.plate_no
    main_dev = await db.scalar(
        select(VehicleDevice.device_no)
        .where(VehicleDevice.vehicle_id == vehicle_id, VehicleDevice.is_main.is_(True))
        .limit(1)
    )
    if not main_dev:
        main_dev = await db.scalar(
            select(VehicleDevice.device_no).where(VehicleDevice.vehicle_id == vehicle_id).limit(1)
        )
    await db.execute(delete(VehicleAllocRuleVehicle).where(VehicleAllocRuleVehicle.vehicle_id == vehicle_id))
    await db.execute(delete(VehicleLocation).where(VehicleLocation.vehicle_id == vehicle_id))
    await db.execute(delete(VehicleDevice).where(VehicleDevice.vehicle_id == vehicle_id))
    await db.execute(delete(Vehicle).where(Vehicle.id == vehicle_id))
    await db.commit()
    jt = await jt808_vehicle.delete_now(main_dev, plate_no) if (main_dev or plate_no) else None
    return plate_no, ("ok" if jt else ("fail" if jt is False else "skip"))


async def run(xlsx: Path, apply: bool) -> int:
    from app.database import AsyncSessionLocal
    from app.models import Fleet, OrgCompany, Vehicle, VehicleDevice
    from app.routers.api_vehicle import _stored_device_no_for_status
    from app import jt808_vehicle

    rows = load_excel_rows(xlsx)
    logger.info("Excel 有效车牌 %s", len(rows))

    async with AsyncSessionLocal() as db:
        companies = (await db.execute(select(OrgCompany).order_by(OrgCompany.id))).scalars().all()
        by_name: dict[str, list[CompanyRef]] = {}
        for c in companies:
            ref = CompanyRef(id=int(c.id), name=_norm(c.name), org_code=_norm(c.org_code))
            by_name.setdefault(ref.name, []).append(ref)

        fleets = (await db.execute(select(Fleet).order_by(Fleet.id))).scalars().all()
        fleet_by_co: dict[int, dict[str, int]] = {}
        for fl in fleets:
            fleet_by_co.setdefault(int(fl.company_id), {})[_norm(fl.name)] = int(fl.id)

        vehicles = (await db.execute(select(Vehicle))).scalars().all()
        by_plate = {_norm_plate(v.plate_no): v for v in vehicles}
        vid_to_vehicle = {int(v.id): v for v in vehicles}
        devices = (await db.execute(select(VehicleDevice))).scalars().all()
        dev_by_vid: dict[int, Any] = {}
        vid_by_dev: dict[str, int] = {}
        for d in devices:
            if d.is_main:
                dev_by_vid[int(d.vehicle_id)] = d
            if _norm(d.device_no):
                vid_by_dev[_norm(d.device_no)] = int(d.vehicle_id)
        excel_dev_by_plate = {
            rec["plate"]: rec["device"] for rec in rows if rec.get("device")
        }

        def resolve_company(name: str, fleet_name: str) -> tuple[CompanyRef | None, str]:
            # 系统里车队也是 org_company；Excel「车队」优先，「所属公司」仅兜底。
            for key in (fleet_name, name):
                key = _norm(key)
                if not key:
                    continue
                hits = by_name.get(key) or []
                if len(hits) == 1:
                    return hits[0], key
                if len(hits) > 1:
                    return None, f"公司重名无法唯一匹配：{key}"
            return None, f"公司信息中无此名称：{name or fleet_name}"

        async def free_holder(store_no: str, keep_vid: int) -> None:
            holder = vid_by_dev.get(store_no)
            if holder is None or holder == keep_vid:
                return
            other_d = dev_by_vid.get(int(holder))
            if other_d is None:
                other_d = await db.scalar(
                    select(VehicleDevice).where(VehicleDevice.device_no == store_no).limit(1)
                )
            if other_d is None:
                return
            tmp = f"FREED{holder}"
            n = 0
            while tmp in vid_by_dev:
                n += 1
                tmp = f"FREED{holder}_{n}"
            vid_by_dev.pop(_norm(other_d.device_no), None)
            other_d.device_no = tmp
            vid_by_dev[tmp] = int(holder)
            await db.flush()
            logger.info("释放设备号 %s <- vehicle_id=%s 临时号 %s", store_no, holder, tmp)

        report = {
            "excel": len(rows),
            "cesg": len(vehicles),
            "update": [],
            "create": [],
            "delete": [],
            "skip": [],
            "company_miss": [],
            "fleet_miss": [],
            "sync": {"ok": 0, "fail": 0, "skip": 0},
        }
        sync_queue: list[tuple[int, str | None]] = []
        excel_plates: set[str] = {rec["plate"] for rec in rows}
        extra = [v for p, v in by_plate.items() if p not in excel_plates]
        for v in extra:
            report["delete"].append(
                {
                    "id": v.id,
                    "plate": v.plate_no,
                    "status": v.status,
                    "company_id": v.company_id,
                }
            )
        if apply and extra:
            deleted_early = []
            for v in extra:
                plate, st = await _delete_cesg_vehicle(db, int(v.id))
                deleted_early.append({"plate": plate or v.plate_no, "jt808": st})
                by_plate.pop(_norm_plate(v.plate_no), None)
                vid_to_vehicle.pop(int(v.id), None)
                old_d = dev_by_vid.pop(int(v.id), None)
                if old_d is not None:
                    vid_by_dev.pop(_norm(old_d.device_no), None)
            report["deleted"] = deleted_early
            logger.info("已先删除 CESG 多余 %s，释放设备号", len(deleted_early))

        for rec in rows:
            company, how = resolve_company(rec["company"], rec["fleet"])
            if company is None:
                report["skip"].append({"plate": rec["plate"], "reason": how, "row": rec["row"]})
                if how not in report["company_miss"]:
                    report["company_miss"].append(how)
                continue
            if not rec["device"]:
                report["skip"].append({"plate": rec["plate"], "reason": "缺少设备编号", "row": rec["row"]})
                continue

            v = by_plate.get(rec["plate"])
            status = (v.status if v is not None else "正常") or "正常"
            store_no = _stored_device_no_for_status(rec["device"], status)
            owner = vid_by_dev.get(store_no)
            if owner is not None and (v is None or int(v.id) != owner):
                owner_v = vid_to_vehicle.get(int(owner))
                owner_plate = _norm_plate(owner_v.plate_no) if owner_v is not None else ""
                owner_excel_dev = excel_dev_by_plate.get(owner_plate, "")
                owner_will_leave = False
                if owner_plate and owner_plate not in excel_plates:
                    owner_will_leave = True
                elif owner_excel_dev:
                    owner_store = _stored_device_no_for_status(
                        owner_excel_dev, (owner_v.status if owner_v is not None else "正常")
                    )
                    owner_will_leave = owner_store != store_no
                if not owner_will_leave:
                    report["skip"].append(
                        {
                            "plate": rec["plate"],
                            "reason": f"设备号已被其它车辆占用 vehicle_id={owner} plate={owner_plate}：{store_no}",
                            "row": rec["row"],
                        }
                    )
                    continue

            fleet_id = None
            fleet_note = ""
            if rec["fleet"]:
                fl_id = fleet_by_co.get(int(company.id), {}).get(rec["fleet"])
                if fl_id is None:
                    fleet_note = f"车队不存在或不属于该公司：{rec['fleet']} / {company.name}"
                    report["fleet_miss"].append(
                        {"plate": rec["plate"], "fleet": rec["fleet"], "company": company.name}
                    )
                else:
                    fleet_id = fl_id

            changes: list[str] = []
            is_new = v is None
            old_dev = _norm(dev_by_vid[int(v.id)].device_no) if (v is not None and int(v.id) in dev_by_vid) else ""

            if apply:
                try:
                    async with db.begin_nested():
                        if is_new:
                            v = Vehicle(plate_no=rec["plate"], status="正常")
                            db.add(v)
                            await db.flush()
                        org_code = company.org_code or f"{company.id:04d}"
                        _set_if(v, "company_id", int(company.id), changes)
                        _set_if(v, "company_org_code", org_code, changes)
                        if fleet_id is not None:
                            _set_if(v, "fleet_id", int(fleet_id), changes)
                        _set_if(v, "vin", rec["vin"] or None, changes)
                        _set_if(v, "frame_no", rec["vin"] or None, changes)
                        _set_if(v, "install_date", rec["install_date"], changes)
                        _set_if(v, "brand", rec["brand"] or None, changes)
                        _set_if(v, "model", rec["model"] or None, changes)
                        _set_if(v, "vehicle_category", rec["category"], changes)
                        if rec["channel"] is not None:
                            _set_if(v, "channel_count", rec["channel"], changes)
                        await db.flush()
                        await free_holder(store_no, int(v.id))
                        d = dev_by_vid.get(int(v.id))
                        if d is None:
                            d = VehicleDevice(
                                vehicle_id=int(v.id),
                                is_main=True,
                                channel_no=1,
                                device_no=store_no,
                                sim_no=rec["sim"] or None,
                            )
                            db.add(d)
                            await db.flush()
                            dev_by_vid[int(v.id)] = d
                            changes.append(f"device_no:new={store_no!r}")
                        elif _norm(d.device_no) != store_no:
                            if old_dev and old_dev in vid_by_dev:
                                vid_by_dev.pop(old_dev, None)
                            d.device_no = store_no
                            changes.append(f"device_no:{old_dev!r}->{store_no!r}")
                        if rec["sim"] and _norm(d.sim_no) != rec["sim"]:
                            changes.append(f"sim_no:{d.sim_no!r}->{rec['sim']!r}")
                            d.sim_no = rec["sim"]
                        vid_by_dev[store_no] = int(v.id)
                    if is_new:
                        by_plate[rec["plate"]] = v
                        vid_to_vehicle[int(v.id)] = v
                    sync_queue.append((int(v.id), old_dev or None))
                    await db.commit()
                except Exception as exc:  # noqa: BLE001
                    report["skip"].append(
                        {"plate": rec["plate"], "reason": f"写入失败：{exc}", "row": rec["row"]}
                    )
                    logger.warning("写入失败 plate=%s: %s", rec["plate"], exc)
                    continue
            else:
                if is_new:
                    changes.append("NEW")
                if v is not None and int(getattr(v, "company_id", 0) or 0) != int(company.id):
                    changes.append(f"company_id:{v.company_id}->{company.id}({company.name})")
                if fleet_id is not None and v is not None and getattr(v, "fleet_id", None) != fleet_id:
                    changes.append(f"fleet_id:{v.fleet_id}->{fleet_id}")
                if rec["vin"] and v is not None and _norm(v.vin) != rec["vin"]:
                    changes.append("vin")
                if rec["device"] and old_dev and old_dev != store_no:
                    changes.append(f"device:{old_dev}->{store_no}")
                elif rec["device"] and not old_dev:
                    changes.append(f"device:new={store_no}")
                if rec["sim"] and v is not None:
                    cur_sim = _norm(dev_by_vid[int(v.id)].sim_no) if int(v.id) in dev_by_vid else ""
                    if cur_sim != rec["sim"]:
                        changes.append("sim")
                if fleet_note:
                    changes.append(fleet_note)

            item = {
                "plate": rec["plate"],
                "company": company.name,
                "match": how,
                "device": store_no,
                "changes": changes,
                "fleet_note": fleet_note,
            }
            if is_new:
                report["create"].append(item)
            else:
                report["update"].append(item)

        if apply:
            await db.commit()
            logger.info("CESG 写入完成 create=%s update=%s skip=%s", len(report["create"]), len(report["update"]), len(report["skip"]))

    if apply and sync_queue:
        total = len(sync_queue)
        for i, (vid, old_dev) in enumerate(sync_queue, start=1):
            try:
                result = await jt808_vehicle.upsert_now(vid, old_dev)
            except Exception as exc:  # noqa: BLE001
                report["sync"]["fail"] += 1
                logger.warning("808 同步异常 vehicle_id=%s: %s", vid, exc)
                continue
            if result is True:
                report["sync"]["ok"] += 1
            elif result is False:
                report["sync"]["fail"] += 1
            else:
                report["sync"]["skip"] += 1
            if i % 50 == 0 or i == total:
                logger.info("808 同步进度 %s/%s ok=%s fail=%s skip=%s", i, total, report["sync"]["ok"], report["sync"]["fail"], report["sync"]["skip"])
        logger.info("808 同步 ok=%s fail=%s skip=%s", report["sync"]["ok"], report["sync"]["fail"], report["sync"]["skip"])

    summary = {
        "mode": "apply" if apply else "dry-run",
        "excel": report["excel"],
        "cesg": report["cesg"],
        "create": len(report["create"]),
        "update": len(report["update"]),
        "delete": len(report["delete"]),
        "skip": len(report["skip"]),
        "company_miss": report["company_miss"],
        "fleet_miss_count": len(report["fleet_miss"]),
        "sync": report["sync"],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("==== SKIP ====")
    for x in report["skip"][:80]:
        print(json.dumps(x, ensure_ascii=False))
    print("==== CREATE ====")
    for x in report["create"][:40]:
        print(json.dumps(x, ensure_ascii=False))
    print("==== DELETE ====")
    for x in report["delete"]:
        print(json.dumps(x, ensure_ascii=False))
    print("==== FLEET MISS ====")
    seen_f = set()
    fleet_unique = []
    for x in report["fleet_miss"]:
        key = (x["company"], x["fleet"])
        if key in seen_f:
            continue
        seen_f.add(key)
        item = {"company": x["company"], "fleet": x["fleet"]}
        fleet_unique.append(item)
        print(json.dumps(item, ensure_ascii=False))
    report_path = Path(xlsx).with_suffix(".reconcile-report.json")
    report_path.write_text(
        json.dumps(
            {
                "summary": summary,
                "skip": report["skip"],
                "create": report["create"],
                "delete": report["delete"],
                "fleet_miss_unique": fleet_unique,
                "deleted": report.get("deleted"),
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )
    print("REPORT", str(report_path))
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="按 CAR.XLSX 对齐 CESG / 808 车辆")
    parser.add_argument("--xlsx", required=True, help="CAR.XLSX 路径")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true")
    group.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        raise SystemExit(f"找不到文件：{xlsx}")
    return asyncio.run(run(xlsx, apply=bool(args.apply)))


if __name__ == "__main__":
    raise SystemExit(main())
