"""车辆相关 API — 与原型 车辆信息/添加车辆 对应。

在线状态：online_source=db 读 Vehicle.is_connect；其余按 last_online_at 空闲窗口判定。
（设备/视频/实时仍由 808 平台负责，本后端不探测设备网关。）
"""
import base64
import os
from datetime import date, datetime, timedelta, timezone

from app.timeutil import china_now_naive
from io import BytesIO

from fastapi import APIRouter, BackgroundTasks, Depends, File, Header, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app import jt808_vehicle
from app.database import get_db
from app.jt808_openapi_client import jt808_openapi_client
from app.models import (
    Driver,
    Fleet,
    OrgCompany,
    ReserveTerminal,
    Vehicle,
    VehicleAllocRuleVehicle,
    VehicleDevice,
)
from app.org_scope import (
    collect_org_company_subtree_ids,
    require_x_org_id_header,
    wants_org_tree_scope,
)
from app.risk_profile_service import (
    expand_company_filter_ids,
    load_org_company_maps,
    resolve_selected_to_local_org_ids,
)
from app.vehicle_alloc_scope import (
    apply_vehicle_id_scope,
    parse_user_id_header,
    resolve_allowed_vehicle_ids,
)

router = APIRouter(prefix="/api/vehicle", tags=["vehicle"])


def _online_time_candidates(*values: str | None) -> list[str]:
    out: list[str] = []
    for raw in values:
        s = str(raw or "").strip()
        if not s or s in out:
            continue
        out.append(s)
        digits = "".join(ch for ch in s if ch.isdigit())
        if digits and digits != s and digits not in out:
            out.append(digits)
        if digits.isdigit() and 0 < len(digits) < 12:
            padded = digits.zfill(12)
            if padded not in out:
                out.append(padded)
    return out


def _format_travel_stime(raw: str) -> str:
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    if len(digits) >= 14:
        return (
            f"{digits[0:4]}-{digits[4:6]}-{digits[6:8]} "
            f"{digits[8:10]}:{digits[10:12]}:{digits[12:14]}"
        )
    return str(raw or "").strip()


@router.get("/online-time")
async def vehicle_online_time(
    tid: str | None = Query(None),
    deviceId: str | None = Query(None),
    car_id: str | None = Query(None),
):
    """实时监控气泡：今日首次行车开始时间（808 无会话上线字段时的近似）。"""
    candidates = _online_time_candidates(tid, deviceId, car_id)
    if not candidates:
        return {"ok": True, "data": {"online_time": None}}
    now = china_now_naive()
    day = now.strftime("%Y%m%d")
    stime = f"{day}000000"
    etime = f"{day}235959"
    best = ""
    used = ""
    for device in candidates:
        try:
            result = await jt808_openapi_client.list_travel_records(
                device_id=device,
                stime=stime,
                etime=etime,
                type=1,
                page=1,
                rows=50,
            )
        except Exception:
            continue
        rows = result.get("data") if isinstance(result.get("data"), list) else []
        for row in rows:
            raw = str((row or {}).get("stime") or (row or {}).get("start_time") or "").strip()
            if raw and (not best or raw < best):
                best = raw
                used = device
        if best:
            break
    text = _format_travel_stime(best)
    return {
        "ok": True,
        "data": {
            "online_time": text or None,
            "raw": best or None,
            "deviceId": used or None,
        },
    }


def _jt808_sync_status(result: bool | None) -> str:
    if result is True:
        return "success"
    if result is False:
        return "failed"
    return "skipped"


def _parent_org_name(
    org_id: int | None,
    company_map: dict[int, str | None],
    parent_map: dict[int, int | None],
) -> str | None:
    parent_id = parent_map.get(org_id) if org_id else None
    return company_map.get(parent_id) if parent_id else None


def _find_ancestor_with_fleet_name(
    org_id: int | None,
    company_map: dict[int, str | None],
    parent_map: dict[int, int | None],
) -> int | None:
    """从 org 的直接上级开始向上，返回第一个名称含“车队”的组织 id。"""
    current_id = parent_map.get(org_id) if org_id else None
    visited: set[int] = set()
    while current_id and current_id not in visited:
        visited.add(current_id)
        name = company_map.get(current_id)
        if name and "车队" in name:
            return current_id
        current_id = parent_map.get(current_id)
    return None


def _is_project_or_group_like_name(name: str | None) -> bool:
    text = _norm(name)
    return bool(text) and ("项目" in text or "组" in text)


def _vehicle_list_company_fleet_names(
    company_id: int | None,
    fleet_id: int | None,
    company_map: dict[int, str | None],
    parent_map: dict[int, int | None],
    fleet_map: dict[int, str | None],
) -> tuple[str | None, str | None]:
    """返回列表展示用的 (所属公司, 车队)。"""
    if not company_id:
        return None, fleet_map.get(fleet_id) if fleet_id else None

    company_name = company_map.get(company_id)

    # 1) 本级名称含“车队”：车队=本级，所属公司=上级
    if company_name and "车队" in company_name:
        return _parent_org_name(company_id, company_map, parent_map), company_name

    # 2) 向上找含“车队”的上级：车队=该上级，所属公司=该上级的上级
    fleet_org_id = _find_ancestor_with_fleet_name(company_id, company_map, parent_map)
    if fleet_org_id is not None:
        return (
            _parent_org_name(fleet_org_id, company_map, parent_map),
            company_map.get(fleet_org_id),
        )

    # 3) 找不到“车队”，且本级名称含“项目/组”等：所属公司=上级，车队为空
    if _is_project_or_group_like_name(company_name):
        return _parent_org_name(company_id, company_map, parent_map), None

    # 4) 其它情况维持原逻辑
    return company_name, fleet_map.get(fleet_id) if fleet_id else None


def _org_level_names(
    org_id: int | None,
    company_map: dict[int, str | None],
    parent_map: dict[int, int | None],
    max_levels: int = 4,
) -> list[str | None]:
    """车辆所属组织沿 parent_id 上溯到根，映射为一至四级名称。

    一级一般为集团总公司（根节点）；不足四级的位置返回 None（前端显示 --）。
    超过四级时：前三级各占一位，第四位为第 4 级及以下名称用 "/" 拼接。
    """
    if not org_id:
        return [None] * max_levels
    chain: list[int] = []
    current: int | None = int(org_id)
    visited: set[int] = set()
    while current and current not in visited:
        visited.add(current)
        chain.append(current)
        current = parent_map.get(current)
    chain.reverse()
    full_names: list[str] = []
    for cid in chain:
        name = company_map.get(cid)
        text = (name or "").strip() if isinstance(name, str) else ""
        if text:
            full_names.append(text)
    names: list[str | None] = [
        full_names[0] if len(full_names) > 0 else None,
        full_names[1] if len(full_names) > 1 else None,
        full_names[2] if len(full_names) > 2 else None,
        ("/".join(full_names[3:]) if len(full_names) > 3 else None),
    ]
    # 保持返回长度与历史一致（固定 4）
    while len(names) < max_levels:
        names.append(None)
    return names[:max_levels]


def _norm(s) -> str:
    if s is None:
        return ""
    return str(s).replace("\u3000", " ").strip()


def _terminal_variants(terminal_id: str) -> list[str]:
    t = (terminal_id or "").strip()
    if not t:
        return []
    variants = {t}
    stripped = t.lstrip("0")
    if stripped:
        variants.add(stripped)
    if t.isdigit():
        variants.add(t.zfill(12))
    return list(variants)


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = _norm(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _to_datetime(v) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = _norm(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _to_int(v, default: int | None = None) -> int | None:
    s = _norm(v)
    if not s:
        return default
    try:
        return int(float(s))
    except Exception:
        return default


def _to_float(v, default: float | None = None) -> float | None:
    s = _norm(v)
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        return default


def _to_bool(v, default: bool = False) -> bool:
    if isinstance(v, bool):
        return v
    s = _norm(v).lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    return default


def _online_idle_sec() -> float:
    try:
        raw = float((os.getenv("JT808_ONLINE_IDLE_SEC") or "90").strip() or "90")
    except Exception:
        raw = 90.0
    return max(5.0, min(raw, 86400.0))


def _is_vehicle_online(last_online_at: datetime | None) -> bool:
    """库里 naive datetime 统一按东八区解释（与 china_now_naive 写入约定一致）。"""
    if last_online_at is None:
        return False
    if last_online_at.tzinfo:
        la = last_online_at.astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)
    else:
        la = last_online_at
    return (china_now_naive() - la).total_seconds() <= _online_idle_sec()


class VehicleSavePayload(BaseModel):
    vehicle_id: int | None = None
    plate_no: str = Field(..., min_length=1, max_length=16)
    plate_color: str | None = None
    vehicle_category: str | None = None
    vehicle_type: str | None = None
    vehicle_type_ii: str | None = None
    color: str | None = None
    vin: str | None = None
    driving_license_no: str | None = None
    engine_no: str | None = None
    product_model_code: str | None = None
    frame_no: str | None = None
    vehicle_type_code: str | None = None
    vehicle_length: float | None = None
    vehicle_width: float | None = None
    vehicle_height: float | None = None
    loaded_weight: float | None = None
    vehicle_payload: float | None = None
    curb_weight: float | None = None
    urea_info: str | None = None
    coolant_temp_high_threshold: float | None = None
    fuel_level_low_threshold: float | None = None
    reagent_level_low_threshold: float | None = None
    dpf_pressure_high_threshold: float | None = None
    scr_downstream_abnormal_threshold: float | None = None
    nox_abnormal_threshold: float | None = None
    short_name: str | None = None
    company_id: int = Field(..., ge=1)
    fleet_id: int | None = None
    driver_id: int | None = None
    driver_name: str | None = None
    owner_name: str | None = None
    contact_name: str | None = None
    contact_phone: str | None = None
    legal_contact_phone: str | None = None
    legal_address: str | None = None
    route: str | None = None
    agent: str | None = None
    install_date: str | None = None
    service_start_date: str | None = None
    service_end_date: str | None = None
    status: str | None = None
    last_online_at: str | None = None
    channel_count: int | None = None
    engine_displacement: str | None = None
    fuel_tank_capacity: str | None = None
    fuel_tank: str | None = None
    rapid_acceleration: str | None = None
    rapid_deceleration: str | None = None
    battery_capacity: str | None = None
    range_mileage: str | None = None
    battery_no: str | None = None
    motor_no: str | None = None
    manufacturer: str | None = None
    brand: str | None = None
    model: str | None = None
    vehicle_grade: str | None = None
    vehicle_usage: str | None = None
    speed_limit: float | None = None
    track_retain_days: int | None = None
    mileage_factor: float | None = None
    mileage_offset: float | None = None
    scrap_date: str | None = None
    inspect_date: str | None = None
    plate_login: bool | None = None
    is_connect: bool | None = None
    night_speed_enabled: bool | None = None
    night_start_time: str | None = None
    night_end_time: str | None = None
    night_speed_percent: float | None = None
    icon_id: int | None = None
    remark: str | None = None
    device_no: str | None = None
    device_sn: str | None = None
    terminal_type: str | None = None
    sim_no: str | None = None
    actual_sim: str | None = None
    product_model: str | None = None
    channels: list[str] | None = None
    device_channel_no: int | None = None


async def _ensure_driver_binding(db: AsyncSession, driver_id: int | None, company_id: int) -> None:
    if driver_id is None or driver_id <= 0:
        return
    dr = await db.scalar(select(Driver).where(Driver.id == driver_id).limit(1))
    if dr is None:
        raise HTTPException(status_code=400, detail="绑定司机不存在")
    if dr.company_id != company_id:
        raise HTTPException(status_code=400, detail="绑定司机不属于所选公司")


async def _ensure_company_and_fleet(db: AsyncSession, company_id: int, fleet_id: int | None) -> OrgCompany:
    co = await db.scalar(select(OrgCompany).where(OrgCompany.id == company_id).limit(1))
    if co is None:
        raise HTTPException(status_code=400, detail="所属公司不存在")
    if fleet_id:
        fl = await db.scalar(select(Fleet).where(Fleet.id == fleet_id).limit(1))
        if fl is None or fl.company_id != company_id:
            raise HTTPException(status_code=400, detail="车队不存在或不属于所属公司")
    return co


async def _remove_reserve_for_device_no(db: AsyncSession, device_no: str) -> None:
    t = _norm(device_no)
    if not t:
        return
    ids = _terminal_variants(t)
    if ids:
        await db.execute(delete(ReserveTerminal).where(ReserveTerminal.terminal_id.in_(ids)))


async def _upsert_main_device(db: AsyncSession, vehicle_id: int, payload: VehicleSavePayload) -> None:
    dev_no = _norm(payload.device_no)
    dr = await db.execute(
        select(VehicleDevice).where(VehicleDevice.vehicle_id == vehicle_id, VehicleDevice.is_main.is_(True))
    )
    d = dr.scalar_one_or_none()
    if not dev_no:
        if d is not None:
            await db.execute(delete(VehicleDevice).where(VehicleDevice.id == d.id))
        return
    if d is None:
        d = VehicleDevice(vehicle_id=vehicle_id, is_main=True)
        db.add(d)
    d.device_no = dev_no
    d.device_sn = _norm(payload.device_sn) or None
    d.terminal_type = _norm(payload.terminal_type) or None
    d.sim_no = _norm(payload.sim_no) or None
    d.actual_sim = _norm(payload.actual_sim) or None
    d.product_model = _norm(payload.product_model) or None
    d.channels = payload.channels or []
    d.channel_no = _to_int(payload.device_channel_no, 1) or 1


async def _ensure_unique_plate_and_device(
    db: AsyncSession, payload: VehicleSavePayload, current_vehicle_id: int | None = None
) -> None:
    plate = _norm(payload.plate_no)
    if not plate:
        raise HTTPException(status_code=400, detail="车牌号不能为空")
    q_plate = select(Vehicle.id).where(Vehicle.plate_no == plate)
    if current_vehicle_id:
        q_plate = q_plate.where(Vehicle.id != current_vehicle_id)
    if await db.scalar(q_plate.limit(1)) is not None:
        raise HTTPException(status_code=400, detail="车牌号已存在")

    device_no = _norm(payload.device_no)
    if not device_no:
        return
    q_dev = select(VehicleDevice.vehicle_id).where(VehicleDevice.device_no == device_no)
    if current_vehicle_id:
        q_dev = q_dev.where(VehicleDevice.vehicle_id != current_vehicle_id)
    if await db.scalar(q_dev.limit(1)) is not None:
        raise HTTPException(status_code=400, detail="设备号已存在")


def _apply_vehicle_payload(v: Vehicle, co: OrgCompany, payload: VehicleSavePayload) -> None:
    v.plate_no = _norm(payload.plate_no)
    v.plate_color = _norm(payload.plate_color) or "黄牌"
    v.vehicle_category = _norm(payload.vehicle_category) or None
    v.vehicle_type = _norm(payload.vehicle_type) or None
    v.vehicle_type_ii = _norm(payload.vehicle_type_ii) or None
    v.color = _norm(payload.color) or None
    v.vin = _norm(payload.vin) or None
    v.driving_license_no = _norm(payload.driving_license_no) or None
    v.engine_no = _norm(payload.engine_no) or None
    v.product_model_code = _norm(payload.product_model_code) or None
    v.frame_no = _norm(payload.frame_no) or None
    v.vehicle_type_code = _norm(payload.vehicle_type_code) or None
    v.vehicle_length = _to_float(payload.vehicle_length, None)
    v.vehicle_width = _to_float(payload.vehicle_width, None)
    v.vehicle_height = _to_float(payload.vehicle_height, None)
    v.loaded_weight = _to_float(payload.loaded_weight, None)
    v.vehicle_payload = _to_float(payload.vehicle_payload, None)
    v.curb_weight = _to_float(payload.curb_weight, None)
    v.urea_info = _norm(payload.urea_info) or None
    v.coolant_temp_high_threshold = _to_float(payload.coolant_temp_high_threshold, None)
    v.fuel_level_low_threshold = _to_float(payload.fuel_level_low_threshold, None)
    v.reagent_level_low_threshold = _to_float(payload.reagent_level_low_threshold, None)
    v.dpf_pressure_high_threshold = _to_float(payload.dpf_pressure_high_threshold, None)
    v.scr_downstream_abnormal_threshold = _to_float(payload.scr_downstream_abnormal_threshold, None)
    v.nox_abnormal_threshold = _to_float(payload.nox_abnormal_threshold, None)
    v.short_name = _norm(payload.short_name) or None
    v.company_id = payload.company_id
    v.company_org_code = _norm(co.org_code) or None
    v.fleet_id = payload.fleet_id
    v.driver_id = payload.driver_id if payload.driver_id and payload.driver_id > 0 else None
    v.driver_name = _norm(payload.driver_name) or None
    v.owner_name = _norm(payload.owner_name) or None
    v.contact_name = _norm(payload.contact_name) or None
    v.contact_phone = _norm(payload.contact_phone) or None
    v.legal_contact_phone = _norm(payload.legal_contact_phone) or None
    v.legal_address = _norm(payload.legal_address) or None
    v.route = _norm(payload.route) or None
    v.agent = _norm(payload.agent) or None
    v.install_date = _to_date(payload.install_date)
    v.service_start_date = _to_date(payload.service_start_date)
    v.service_end_date = _to_date(payload.service_end_date)
    v.status = _norm(payload.status) or "正常"
    v.last_online_at = _to_datetime(payload.last_online_at)
    channels = payload.channels or []
    cc = _to_int(payload.channel_count, 0) or 0
    v.channel_count = max(cc, len(channels)) if channels else cc
    v.engine_displacement = _norm(payload.engine_displacement) or None
    v.fuel_tank_capacity = _norm(payload.fuel_tank_capacity) or None
    v.fuel_tank = _norm(payload.fuel_tank) or None
    v.rapid_acceleration = _norm(payload.rapid_acceleration) or None
    v.rapid_deceleration = _norm(payload.rapid_deceleration) or None
    v.battery_capacity = _norm(payload.battery_capacity) or None
    v.range_mileage = _norm(payload.range_mileage) or None
    v.battery_no = _norm(payload.battery_no) or None
    v.motor_no = _norm(payload.motor_no) or None
    v.manufacturer = _norm(payload.manufacturer) or None
    v.brand = _norm(payload.brand) or None
    v.model = _norm(payload.model) or None
    v.vehicle_grade = _norm(payload.vehicle_grade) or None
    v.vehicle_usage = _norm(payload.vehicle_usage) or None
    v.speed_limit = _to_float(payload.speed_limit, 0.0) or 0.0
    v.track_retain_days = _to_int(payload.track_retain_days, 0) or 0
    v.mileage_factor = _to_float(payload.mileage_factor, None)
    v.mileage_offset = _to_float(payload.mileage_offset, None)
    v.scrap_date = _to_date(payload.scrap_date)
    v.inspect_date = _to_date(payload.inspect_date)
    v.plate_login = _to_bool(payload.plate_login, False)
    v.is_connect = _to_bool(payload.is_connect, False)
    v.night_speed_enabled = _to_bool(payload.night_speed_enabled, False)
    v.night_start_time = _norm(payload.night_start_time) or None
    v.night_end_time = _norm(payload.night_end_time) or None
    v.night_speed_percent = _to_float(payload.night_speed_percent, None)
    v.icon_id = _to_int(payload.icon_id, 1) or 1
    v.remark = _norm(payload.remark) or None


@router.post("/create")
async def vehicle_create(payload: VehicleSavePayload, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    co = await _ensure_company_and_fleet(db, payload.company_id, payload.fleet_id)
    await _ensure_driver_binding(db, payload.driver_id if payload.driver_id and payload.driver_id > 0 else None, payload.company_id)
    await _ensure_unique_plate_and_device(db, payload, None)
    v = Vehicle()
    _apply_vehicle_payload(v, co, payload)
    db.add(v)
    await db.flush()
    await _upsert_main_device(db, v.id, payload)
    await _remove_reserve_for_device_no(db, _norm(payload.device_no or ""))
    await db.flush()
    await db.commit()
    jt808_result = await jt808_vehicle.upsert_now(v.id)
    return {"ok": True, "message": "已创建", "data": {"id": v.id}, "jt808_sync": _jt808_sync_status(jt808_result)}


@router.post("/update")
async def vehicle_update(payload: VehicleSavePayload, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    vid = int(payload.vehicle_id or 0)
    if vid <= 0:
        raise HTTPException(status_code=400, detail="vehicle_id 必填")
    v = await db.scalar(select(Vehicle).where(Vehicle.id == vid).limit(1))
    if v is None:
        raise HTTPException(status_code=404, detail="车辆不存在")
    old_dev = await db.scalar(
        select(VehicleDevice.device_no).where(VehicleDevice.vehicle_id == vid, VehicleDevice.is_main.is_(True)).limit(1)
    )
    co = await _ensure_company_and_fleet(db, payload.company_id, payload.fleet_id)
    await _ensure_driver_binding(db, payload.driver_id if payload.driver_id and payload.driver_id > 0 else None, payload.company_id)
    await _ensure_unique_plate_and_device(db, payload, vid)
    _apply_vehicle_payload(v, co, payload)
    await _upsert_main_device(db, vid, payload)
    await _remove_reserve_for_device_no(db, _norm(payload.device_no or ""))
    await db.flush()
    await db.commit()
    jt808_result = await jt808_vehicle.upsert_now(vid, old_dev)
    return {"ok": True, "message": "已更新", "data": {"id": vid}, "jt808_sync": _jt808_sync_status(jt808_result)}


@router.get("/detail/{vehicle_id}")
async def vehicle_detail(
    vehicle_id: int,
    x_user_id: str | None = Header(None, alias="X-User-Id"),
    db: AsyncSession = Depends(get_db),
):
    allowed = await resolve_allowed_vehicle_ids(db, parse_user_id_header(x_user_id))
    if allowed is not None and vehicle_id not in allowed:
        raise HTTPException(status_code=404, detail="车辆不存在")
    v = await db.scalar(select(Vehicle).where(Vehicle.id == vehicle_id).limit(1))
    if v is None:
        raise HTTPException(status_code=404, detail="车辆不存在")
    d = await db.scalar(
        select(VehicleDevice).where(VehicleDevice.vehicle_id == vehicle_id, VehicleDevice.is_main.is_(True)).limit(1)
    )
    driver_name = None
    if v.driver_id:
        driver_name = await db.scalar(select(Driver.name).where(Driver.id == v.driver_id).limit(1))
    return {
        "ok": True,
        "data": {
            "id": v.id,
            "plate_no": v.plate_no,
            "plate_color": v.plate_color,
            "vehicle_category": v.vehicle_category,
            "vehicle_type": v.vehicle_type,
            "vehicle_type_ii": v.vehicle_type_ii,
            "color": v.color,
            "vin": v.vin,
            "driving_license_no": v.driving_license_no,
            "engine_no": v.engine_no,
            "product_model_code": v.product_model_code,
            "frame_no": v.frame_no,
            "vehicle_type_code": v.vehicle_type_code,
            "vehicle_length": float(v.vehicle_length) if v.vehicle_length is not None else None,
            "vehicle_width": float(v.vehicle_width) if v.vehicle_width is not None else None,
            "vehicle_height": float(v.vehicle_height) if v.vehicle_height is not None else None,
            "loaded_weight": float(v.loaded_weight) if v.loaded_weight is not None else None,
            "vehicle_payload": float(v.vehicle_payload) if v.vehicle_payload is not None else None,
            "curb_weight": float(v.curb_weight) if v.curb_weight is not None else None,
            "urea_info": v.urea_info,
            "coolant_temp_high_threshold": (
                float(v.coolant_temp_high_threshold) if v.coolant_temp_high_threshold is not None else None
            ),
            "fuel_level_low_threshold": (
                float(v.fuel_level_low_threshold) if v.fuel_level_low_threshold is not None else None
            ),
            "reagent_level_low_threshold": (
                float(v.reagent_level_low_threshold) if v.reagent_level_low_threshold is not None else None
            ),
            "dpf_pressure_high_threshold": (
                float(v.dpf_pressure_high_threshold) if v.dpf_pressure_high_threshold is not None else None
            ),
            "scr_downstream_abnormal_threshold": (
                float(v.scr_downstream_abnormal_threshold)
                if v.scr_downstream_abnormal_threshold is not None
                else None
            ),
            "nox_abnormal_threshold": (
                float(v.nox_abnormal_threshold) if v.nox_abnormal_threshold is not None else None
            ),
            "short_name": v.short_name,
            "company_id": v.company_id,
            "fleet_id": v.fleet_id,
            "driver_id": v.driver_id,
            "driver_name": v.driver_name or driver_name,
            "owner_name": v.owner_name,
            "contact_name": v.contact_name,
            "contact_phone": v.contact_phone,
            "legal_contact_phone": v.legal_contact_phone,
            "legal_address": v.legal_address,
            "route": v.route,
            "agent": v.agent,
            "install_date": str(v.install_date) if v.install_date else None,
            "service_start_date": str(v.service_start_date) if v.service_start_date else None,
            "service_end_date": str(v.service_end_date) if v.service_end_date else None,
            "status": v.status,
            "last_online_at": v.last_online_at.isoformat() if v.last_online_at else None,
            "channel_count": v.channel_count,
            "engine_displacement": v.engine_displacement,
            "fuel_tank_capacity": v.fuel_tank_capacity,
            "fuel_tank": v.fuel_tank,
            "rapid_acceleration": v.rapid_acceleration,
            "rapid_deceleration": v.rapid_deceleration,
            "battery_capacity": v.battery_capacity,
            "range_mileage": v.range_mileage,
            "battery_no": v.battery_no,
            "motor_no": v.motor_no,
            "manufacturer": v.manufacturer,
            "brand": v.brand,
            "model": v.model,
            "vehicle_grade": v.vehicle_grade,
            "vehicle_usage": v.vehicle_usage,
            "speed_limit": float(v.speed_limit) if v.speed_limit is not None else None,
            "track_retain_days": v.track_retain_days,
            "mileage_factor": float(v.mileage_factor) if v.mileage_factor is not None else None,
            "mileage_offset": float(v.mileage_offset) if v.mileage_offset is not None else None,
            "scrap_date": str(v.scrap_date) if v.scrap_date else None,
            "inspect_date": str(v.inspect_date) if v.inspect_date else None,
            "plate_login": bool(v.plate_login),
            "is_connect": bool(v.is_connect),
            "acc_on": bool(v.acc_on),
            "night_speed_enabled": bool(v.night_speed_enabled),
            "night_start_time": v.night_start_time,
            "night_end_time": v.night_end_time,
            "night_speed_percent": float(v.night_speed_percent) if v.night_speed_percent is not None else None,
            "icon_id": v.icon_id,
            "remark": v.remark,
            "device_no": d.device_no if d else None,
            "device_sn": d.device_sn if d else None,
            "terminal_type": d.terminal_type if d else None,
            "sim_no": d.sim_no if d else None,
            "actual_sim": d.actual_sim if d else None,
            "product_model": d.product_model if d else None,
            "channels": d.channels if d and d.channels else [],
            "device_channel_no": d.channel_no if d else 1,
        },
    }


# 车辆批量导入模板表头（与单车新增必填/808同步字段对齐）
_VEHICLE_IMPORT_HEADERS = [
    "车牌号",
    "车牌颜色",
    "车辆识别代码VIN",
    "车辆发动机号",
    "车辆产品型号代码",
    "车架号",
    "所属公司",
    "车队",
    "设备号",
    "设备类型",
    "SIM卡号",
    "车辆类别",
    "车辆品牌",
    "车辆型号",
    "使用状态",
    "服务开始日",
    "服务到期日",
    "安装日期",
    "通道数",
]


def _vehicle_import_cell(row: tuple, idx: int | None):
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _norm_vehicle_category(raw) -> str | None:
    t = _norm(raw)
    if not t:
        return None
    low = t.lower()
    if low in ("fuel", "燃油", "燃油车"):
        return "fuel"
    if low in ("new", "新能源", "新能源车", "ev", "电动", "电动车"):
        return "new"
    return t


def _header_index(headers: list[str], *names: str, required: bool = True) -> int | None:
    for n in names:
        if n in headers:
            return headers.index(n)
    if required:
        raise HTTPException(
            status_code=400,
            detail=f"缺少表头（需与模板一致，下列其一）：{', '.join(names)}",
        )
    return None


@router.get("/import-template-carinfos")
async def download_vehicle_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "车辆信息导入"
    ws.append(list(_VEHICLE_IMPORT_HEADERS))
    # 示例行（导入前请删除或改成真实数据）
    ws.append(
        [
            "测A12345",
            "黄牌",
            "LSVXXXXXXXXXXXXXXX",
            "ENG001",
            "PMC001",
            "FRAME001",
            "请填写系统中已有公司全称",
            "",
            "123456789012",
            "JT808",
            "",
            "燃油车",
            "",
            "",
            "正常",
            "",
            "",
            "",
            "4",
        ]
    )
    tip = wb.create_sheet("填写说明")
    tip.append(["说明"])
    for line in (
        "1. 请使用本模板填写后上传 .xlsx；带*为导入必填，与页面「添加车辆」一致。",
        "2. 必填：车牌号、车辆识别代码VIN、车辆发动机号、车辆产品型号代码、所属公司、设备号。",
        "3. 所属公司必须与「组织架构」中公司名称完全一致；该公司需已配置 JT808 分组才能同步到 808。",
        "4. 车队可空；填写时须为该公司下已存在的车队名称。",
        "5. 车辆类别可填：燃油车/新能源车（或 fuel/new）。",
        "6. 车牌已存在则更新该车；设备号不可被其它车辆占用。",
        "7. 导入成功后会自动同步到 808 平台（与单车保存相同逻辑）。",
        "8. 填写完请删除本说明页中的示例数据行，或整行改成真实车辆后再导入。",
    ):
        tip.append([line])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    cd = "attachment; filename*=UTF-8''%E8%BD%A6%E8%BE%86%E4%BF%A1%E6%81%AF%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


@router.post("/import-carinfos")
async def import_vehicle_from_carinfos(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件。")
    payload = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(payload), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="Excel 文件无法读取。") from e
    ws = wb.active
    headers = [_norm(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]

    i_plate = _header_index(headers, "车牌号")
    i_plate_color = _header_index(headers, "车牌颜色", required=False)
    i_vin = _header_index(headers, "车辆识别代码VIN", "VIN", "车辆识别代码", required=False)
    i_engine = _header_index(headers, "车辆发动机号", "发动机号", required=False)
    i_pmc = _header_index(headers, "车辆产品型号代码", "产品型号代码", required=False)
    i_frame = _header_index(headers, "车架号", required=False)
    # 兼容旧模板：无「所属公司」时用「车队」列当作公司名
    i_company = _header_index(headers, "所属公司", required=False)
    i_fleet = _header_index(headers, "车队", required=False)
    if i_company is None and i_fleet is None:
        raise HTTPException(status_code=400, detail="缺少表头：所属公司")
    i_start = _header_index(headers, "服务开始日", required=False)
    i_end = _header_index(headers, "服务到期日", required=False)
    i_install = _header_index(headers, "安装日期", required=False)
    i_brand = _header_index(headers, "车辆品牌", required=False)
    i_model = _header_index(headers, "车辆型号", required=False)
    i_category = _header_index(headers, "车辆类别", required=False)
    i_status = _header_index(headers, "使用状态", required=False)
    i_channel = _header_index(headers, "通道数", "通道数目", required=False)
    i_dev_no = _header_index(headers, "设备号", "设备1编号")
    i_dev_type = _header_index(headers, "设备类型", "设备1类型", required=False)
    i_dev_sim = _header_index(headers, "SIM卡号", "设备1SIM", required=False)

    # 旧模板把 VIN 写在「车架号」列：无 VIN 列时回退用车架号列作为 vin
    legacy_vin_from_frame = i_vin is None and i_frame is not None
    if i_vin is None and not legacy_vin_from_frame:
        raise HTTPException(status_code=400, detail="缺少表头：车辆识别代码VIN")

    companies = (await db.execute(select(OrgCompany).order_by(OrgCompany.id))).scalars().all()
    company_name_map: dict[str, list[OrgCompany]] = {}
    for c in companies:
        company_name_map.setdefault((c.name or "").strip(), []).append(c)

    fleets = (await db.execute(select(Fleet).order_by(Fleet.id))).scalars().all()
    fleet_by_company: dict[int, dict[str, Fleet]] = {}
    for fl in fleets:
        fleet_by_company.setdefault(int(fl.company_id), {})[(fl.name or "").strip()] = fl

    def resolve_company(name_key: str) -> OrgCompany:
        cs = company_name_map.get(name_key, [])
        if len(cs) == 1:
            return cs[0]
        if len(cs) > 1:
            raise ValueError(f"公司名称重名无法唯一匹配：{name_key}")
        raise ValueError(f"公司信息中无此名称：{name_key}")

    errors: list[str] = []
    skipped_rows: list[tuple[int, str]] = []
    imported = 0
    updated = 0
    # (vehicle_id, old_device_no|None)
    sync_queue: list[tuple[int, str | None]] = []
    seen_plates: set[str] = set()
    seen_devices: set[str] = set()

    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        plate = _norm(_vehicle_import_cell(row, i_plate))
        company_name = _norm(_vehicle_import_cell(row, i_company)) if i_company is not None else ""
        fleet_name = _norm(_vehicle_import_cell(row, i_fleet)) if i_fleet is not None else ""
        # 旧模板：所属公司空、车队列实际填公司名
        if not company_name and fleet_name and i_company is None:
            company_name = fleet_name
            fleet_name = ""

        if not plate and not company_name and not _norm(_vehicle_import_cell(row, i_dev_no)):
            continue

        def skip(reason: str) -> None:
            errors.append(f"第 {ridx} 行{reason}")
            skipped_rows.append((ridx, reason))

        if not plate:
            skip("缺少车牌号")
            continue
        if plate in seen_plates:
            skip(f"Excel 内车牌号重复：{plate}")
            continue
        if not company_name:
            skip("缺少所属公司")
            continue

        vin = _norm(_vehicle_import_cell(row, i_vin)) if i_vin is not None else ""
        if legacy_vin_from_frame and not vin:
            vin = _norm(_vehicle_import_cell(row, i_frame))
        frame_no = _norm(_vehicle_import_cell(row, i_frame)) if i_frame is not None else ""
        if legacy_vin_from_frame:
            frame_no = ""
        engine_no = _norm(_vehicle_import_cell(row, i_engine)) if i_engine is not None else ""
        product_model_code = _norm(_vehicle_import_cell(row, i_pmc)) if i_pmc is not None else ""
        dev_no = _norm(_vehicle_import_cell(row, i_dev_no))

        if not vin:
            skip("缺少车辆识别代码VIN")
            continue
        if not engine_no:
            skip("缺少车辆发动机号")
            continue
        if not product_model_code:
            skip("缺少车辆产品型号代码")
            continue
        if not dev_no:
            skip("缺少设备号")
            continue
        if dev_no in seen_devices:
            skip(f"Excel 内设备号重复：{dev_no}")
            continue

        try:
            company = resolve_company(company_name)
        except ValueError as e:
            skip(str(e))
            continue

        fleet_id = None
        if fleet_name:
            fl = fleet_by_company.get(int(company.id), {}).get(fleet_name)
            if fl is None:
                skip(f"车队不存在或不属于该公司：{fleet_name}")
                continue
            fleet_id = fl.id

        v = (await db.execute(select(Vehicle).where(Vehicle.plate_no == plate))).scalar_one_or_none()
        old_dev = None
        if v is not None:
            old_dev = await db.scalar(
                select(VehicleDevice.device_no).where(
                    VehicleDevice.vehicle_id == v.id, VehicleDevice.is_main.is_(True)
                ).limit(1)
            )

        dup_q = select(VehicleDevice.vehicle_id).where(VehicleDevice.device_no == dev_no)
        if v is not None and v.id:
            dup_q = dup_q.where(VehicleDevice.vehicle_id != v.id)
        if await db.scalar(dup_q.limit(1)) is not None:
            skip(f"设备号重复：{dev_no}")
            continue

        is_new = v is None
        if is_new:
            v = Vehicle(plate_no=plate)
            db.add(v)
        org_code = _norm(company.org_code) or f"{company.id:04d}"
        v.plate_color = _norm(_vehicle_import_cell(row, i_plate_color)) or "黄牌"
        v.vin = vin
        v.engine_no = engine_no
        v.product_model_code = product_model_code
        v.frame_no = frame_no or None
        v.fleet_id = fleet_id
        v.company_id = company.id
        v.company_org_code = org_code
        v.service_start_date = _to_date(_vehicle_import_cell(row, i_start))
        v.service_end_date = _to_date(_vehicle_import_cell(row, i_end))
        v.install_date = _to_date(_vehicle_import_cell(row, i_install))
        v.brand = _norm(_vehicle_import_cell(row, i_brand)) or None
        v.model = _norm(_vehicle_import_cell(row, i_model)) or None
        v.vehicle_category = _norm_vehicle_category(_vehicle_import_cell(row, i_category))
        v.status = (_norm(_vehicle_import_cell(row, i_status)) or "正常").split(",")[0]
        v.channel_count = _to_int(_vehicle_import_cell(row, i_channel), 0) or 0
        await db.flush()

        d = (
            await db.execute(
                select(VehicleDevice).where(VehicleDevice.vehicle_id == v.id, VehicleDevice.is_main.is_(True))
            )
        ).scalar_one_or_none()
        if d is None:
            d = VehicleDevice(vehicle_id=v.id, is_main=True, channel_no=1)
            db.add(d)
        d.device_no = dev_no
        d.terminal_type = _norm(_vehicle_import_cell(row, i_dev_type)) or None
        d.sim_no = _norm(_vehicle_import_cell(row, i_dev_sim)) or None
        await _remove_reserve_for_device_no(db, dev_no)

        seen_plates.add(plate)
        seen_devices.add(dev_no)
        sync_queue.append((int(v.id), _norm(old_dev) or None))
        if is_new:
            imported += 1
        else:
            updated += 1

    wb.close()
    await db.flush()
    await db.commit()

    jt808_ok = 0
    jt808_fail = 0
    jt808_skip = 0
    for vid, old_dev in sync_queue:
        try:
            result = await jt808_vehicle.upsert_now(vid, old_dev)
        except Exception:  # noqa: BLE001
            jt808_fail += 1
            continue
        status = _jt808_sync_status(result)
        if status == "success":
            jt808_ok += 1
        elif status == "failed":
            jt808_fail += 1
        else:
            jt808_skip += 1

    annotated_file_base64: str | None = None
    annotated_file_name: str | None = None
    if skipped_rows:
        wb_mark = load_workbook(filename=BytesIO(payload))
        ws_mark = wb_mark.active
        max_col = ws_mark.max_column
        mark_col = max_col + 1
        ws_mark.cell(row=1, column=mark_col, value="导入结果")
        bad_fill = PatternFill(fill_type="solid", fgColor="FFF4CCCC")
        for ridx, reason in skipped_rows:
            for c in range(1, max_col + 1):
                ws_mark.cell(row=ridx, column=c).fill = bad_fill
            ws_mark.cell(row=ridx, column=mark_col, value=f"跳过：{reason}")
        bio_mark = BytesIO()
        wb_mark.save(bio_mark)
        wb_mark.close()
        annotated_file_base64 = base64.b64encode(bio_mark.getvalue()).decode("ascii")
        annotated_file_name = "车辆导入结果_含跳过标记.xlsx"

    return {
        "ok": True,
        "imported": imported,
        "updated": updated,
        "skipped": len(skipped_rows),
        "skipped_preview": errors[:20],
        "jt808_sync_ok": jt808_ok,
        "jt808_sync_fail": jt808_fail,
        "jt808_sync_skip": jt808_skip,
        "annotated_file_name": annotated_file_name,
        "annotated_file_base64": annotated_file_base64,
    }


@router.get("/reserve-terminals")
async def vehicle_reserve_terminals(db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(ReserveTerminal).order_by(ReserveTerminal.last_auth_at.desc()))
    rows = r.scalars().all()
    return {
        "ok": True,
        "data": [
            {
                "id": x.id,
                "terminal_id": x.terminal_id,
                "first_auth_at": x.first_auth_at.isoformat() if x.first_auth_at else None,
                "last_auth_at": x.last_auth_at.isoformat() if x.last_auth_at else None,
                "last_peer": x.last_peer,
            }
            for x in rows
        ],
    }


def _parse_company_filter_ids(company_id: int | None, company_ids: str | None) -> list[int]:
    """解析公司筛选：支持本地 org id 或 JT808 gid（与报表公司树一致）。"""
    selected: list[int] = []
    if company_ids:
        for part in str(company_ids).split(","):
            text = part.strip()
            if not text:
                continue
            try:
                selected.append(int(text))
            except (TypeError, ValueError):
                continue
    elif company_id is not None:
        selected.append(int(company_id))
    return selected


@router.get("/list")
async def vehicle_list(
    plate_no: str | None = Query(None),
    status: str | None = Query(None),
    company_id: int | None = Query(None),
    company_ids: str | None = Query(None),
    scope_org_tree: bool = Query(False),
    online_source: str = Query("db"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=1000),
    x_org_id: str | None = Header(None, alias="X-Org-Id"),
    x_user_id: str | None = Header(None, alias="X-User-Id"),
    db: AsyncSession = Depends(get_db),
):
    q = select(Vehicle)
    allowed_vehicle_ids = await resolve_allowed_vehicle_ids(db, parse_user_id_header(x_user_id))
    q = apply_vehicle_id_scope(q, allowed_vehicle_ids)
    if wants_org_tree_scope(scope_org_tree, x_org_id):
        root = require_x_org_id_header(x_org_id)
        co = await db.scalar(select(OrgCompany.id).where(OrgCompany.id == root).limit(1))
        if co is None:
            raise HTTPException(status_code=400, detail="X-Org-Id 对应公司不存在")
        subtree = await collect_org_company_subtree_ids(db, root)
        q = q.where(Vehicle.company_id.in_(subtree))
    if plate_no:
        q = q.where(Vehicle.plate_no.ilike(f"%{plate_no.strip()}%"))
    if status:
        q = q.where(Vehicle.status == status)
    selected_raw = _parse_company_filter_ids(company_id, company_ids)
    if selected_raw:
        parent_map, _, gid_to_local, _ = await load_org_company_maps(db)
        local_selected = resolve_selected_to_local_org_ids(
            selected_raw,
            gid_to_local=gid_to_local,
            known_local_ids=set(parent_map.keys()),
        )
        if local_selected:
            company_subtree = await expand_company_filter_ids(db, local_selected)
            if company_subtree:
                q = q.where(Vehicle.company_id.in_(company_subtree))
    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar() or 0
    # 最新创建的车辆优先，便于列表查看新数据
    q = q.order_by(Vehicle.created_at.desc(), Vehicle.id.desc())
    q = q.offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(q)).scalars().all()

    company_map: dict[int, str | None] = {}
    company_org_map: dict[int, str | None] = {}
    parent_map: dict[int, int | None] = {}
    for cid, cname, ocode, pid in (
        await db.execute(select(OrgCompany.id, OrgCompany.name, OrgCompany.org_code, OrgCompany.parent_id))
    ).all():
        company_map[cid] = cname
        company_org_map[cid] = ocode
        parent_map[cid] = pid

    fleet_map = {}
    for fid, fname in (await db.execute(select(Fleet.id, Fleet.name))).all():
        fleet_map[fid] = fname

    vehicle_ids = [r.id for r in rows]
    main_dev_map: dict[int, VehicleDevice] = {}
    if vehicle_ids:
        dr = await db.execute(
            select(VehicleDevice)
            .where(VehicleDevice.vehicle_id.in_(vehicle_ids))
            .order_by(VehicleDevice.is_main.desc(), VehicleDevice.id.asc())
        )
        for d in dr.scalars().all():
            if d.vehicle_id not in main_dev_map:
                main_dev_map[d.vehicle_id] = d

    driver_ids = sorted({int(r.driver_id) for r in rows if r.driver_id})
    driver_map: dict[int, str] = {}
    if driver_ids:
        for did, dname in (
            await db.execute(select(Driver.id, Driver.name).where(Driver.id.in_(driver_ids)))
        ).all():
            name = (dname or "").strip()
            if name:
                driver_map[int(did)] = name

    osrc = str(online_source or "db").strip().lower()
    use_db_connect = osrc in ("db", "snapshot", "ttx_db", "cache")
    omit_online = osrc in ("omit", "none", "skip", "0", "false")
    items = []
    for r in rows:
        main_dev = main_dev_map.get(r.id)
        if use_db_connect:
            is_online = bool(r.is_connect)
        elif omit_online:
            is_online = False
        else:
            is_online = _is_vehicle_online(r.last_online_at)
        display_company_name, display_fleet_name = _vehicle_list_company_fleet_names(
            r.company_id, r.fleet_id, company_map, parent_map, fleet_map
        )
        level_names = _org_level_names(r.company_id, company_map, parent_map)
        bound_driver_name = (r.driver_name or "").strip() or (
            driver_map.get(int(r.driver_id)) if r.driver_id else None
        )
        items.append(
            {
                "id": r.id,
                "plate_no": r.plate_no,
                "plate_color": r.plate_color,
                "vehicle_type": r.vehicle_type,
                "color": r.color,
                "vin": r.vin,
                "driving_license_no": r.driving_license_no,
                "company_id": r.company_id,
                "company_name": display_company_name,
                "company_org_code": (r.company_org_code or company_org_map.get(r.company_id) or None),
                "fleet_id": r.fleet_id,
                "fleet_name": display_fleet_name,
                "level_one_org": level_names[0],
                "level_two_org": level_names[1],
                "level_three_org": level_names[2],
                "level_four_org": level_names[3],
                "driver_id": r.driver_id,
                "driver_name": bound_driver_name,
                "install_date": str(r.install_date) if r.install_date else None,
                "service_end_date": str(r.service_end_date) if r.service_end_date else None,
                "status": r.status,
                "last_online_at": r.last_online_at.isoformat() if r.last_online_at else None,
                "is_online": is_online,
                "online_status": "在线" if is_online else "离线",
                "contact_name": r.contact_name,
                "contact_phone": r.contact_phone,
                "brand": r.brand,
                "model": r.model,
                "device_no": main_dev.device_no if main_dev else None,
                "device_sn": main_dev.device_sn if main_dev else None,
                "sim_no": main_dev.sim_no if main_dev else None,
                "actual_sim": main_dev.actual_sim if main_dev else None,
                "channel_count": int(r.channel_count) if r.channel_count is not None else 0,
                "acc_on": bool(r.acc_on),
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
        )
    return {"total": total, "items": items, "page": page, "page_size": page_size}


@router.delete("/{vehicle_id}")
async def vehicle_delete(vehicle_id: int, background_tasks: BackgroundTasks, db: AsyncSession = Depends(get_db)):
    row = (await db.execute(select(Vehicle).where(Vehicle.id == vehicle_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="车辆不存在")
    plate_no = row.plate_no
    main_dev = await db.scalar(
        select(VehicleDevice.device_no)
        .where(VehicleDevice.vehicle_id == vehicle_id, VehicleDevice.is_main.is_(True))
        .limit(1)
    )
    if not main_dev:
        main_dev = await db.scalar(
            select(VehicleDevice.device_no).where(VehicleDevice.vehicle_id == vehicle_id).limit(1)
        )
    await db.execute(
        delete(VehicleAllocRuleVehicle).where(VehicleAllocRuleVehicle.vehicle_id == vehicle_id)
    )
    await db.execute(delete(VehicleDevice).where(VehicleDevice.vehicle_id == vehicle_id))
    await db.execute(delete(Vehicle).where(Vehicle.id == vehicle_id))
    await db.commit()
    jt808_result = await jt808_vehicle.delete_now(main_dev, plate_no) if (main_dev or plate_no) else None
    return {"ok": True, "jt808_sync": _jt808_sync_status(jt808_result)}


@router.get("/companies")
async def company_list(db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(select(OrgCompany).order_by(OrgCompany.id))).scalars().all()
    return [{"id": x.id, "name": x.name} for x in rows]


@router.get("/fleets")
async def fleet_list(company_id: int | None = Query(None), db: AsyncSession = Depends(get_db)):
    q = select(Fleet).order_by(Fleet.id)
    if company_id:
        q = q.where(Fleet.company_id == company_id)
    rows = (await db.execute(q)).scalars().all()
    return [{"id": x.id, "name": x.name, "company_id": x.company_id} for x in rows]
