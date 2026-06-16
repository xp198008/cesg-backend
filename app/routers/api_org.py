"""组织架构 / 公司信息 — 共用 org_company 表"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Header, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app import jt808_group
from app.database import get_db
from app.models import Fleet, OrgCompany, Vehicle
from app.org_scope import (
    collect_org_company_subtree_ids,
    require_x_org_id_header,
    wants_org_tree_scope,
)

router = APIRouter(prefix="/api/org", tags=["org"])
logger = logging.getLogger(__name__)


def _gen_org_code(org_id: int) -> str:
    return f"{org_id:04d}"


@router.get("/next-code")
async def next_org_code(db: AsyncSession = Depends(get_db)):
    max_id = await db.scalar(select(func.max(OrgCompany.id)))
    next_id = (max_id or 0) + 1
    return {"next_id": next_id, "org_code": _gen_org_code(next_id)}


class OrgCompanyCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=128)
    legal_person: str | None = Field(None, max_length=64)
    parent_id: int | None = None
    contact_phone: str | None = Field(None, max_length=32)
    address: str | None = Field(None, max_length=256)


class OrgCompanyUpdate(BaseModel):
    name: str | None = Field(default=None, min_length=1, max_length=128)
    legal_person: str | None = Field(default=None, max_length=64)
    parent_id: int | None = None
    contact_phone: str | None = Field(default=None, max_length=32)
    address: str | None = Field(default=None, max_length=256)


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\u3000", " ").strip()


def _norm_parent_cell(v: Any) -> str:
    s = _norm_cell(v)
    if s in {"", "-", "--", "/", "无", "无上级", "根节点", "无（总公司）", "N/A", "n/a", "NA"}:
        return ""
    return s


def _parse_import_rows(file_bytes: bytes) -> list[dict[str, str | None]]:
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="Excel 文件无法读取，请上传有效的 .xlsx 文件。") from e
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_norm_cell(v) for v in (header_cells or ())]

    def _find_header_idx(candidates: tuple[str, ...]) -> int | None:
        for i, h in enumerate(headers):
            if h in candidates:
                return i
        return None

    name_idx = _find_header_idx(("公司名称", "组织名称", "架构名称", "名称"))
    parent_idx = _find_header_idx(("上级公司", "上级单位", "上级组织", "父级公司", "父级单位"))
    if name_idx is None:
        name_idx = 0
    if parent_idx is None:
        parent_idx = 1
    max_col = max(name_idx, parent_idx) + 1

    rows: list[dict[str, str | None]] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2, max_col=max_col, values_only=True), start=2):
        name = _norm_cell(row[name_idx] if len(row) > name_idx else "")
        parent = _norm_parent_cell(row[parent_idx] if len(row) > parent_idx else "")
        if not name and not parent:
            continue
        if not name:
            raise HTTPException(status_code=400, detail=f"第 {ridx} 行公司名称为空。")
        rows.append({"name": name, "parent_name": parent or None})
    wb.close()
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 没有可导入的数据。")
    return rows


def _dedupe_and_validate_import_rows(rows: list[dict[str, str | None]]) -> list[dict[str, str | None]]:
    deduped: list[dict[str, str | None]] = []
    seen_pairs: set[tuple[str, str]] = set()
    all_names_raw = {(r["name"] or "") for r in rows}

    for row in rows:
        name = row["name"] or ""
        parent = row["parent_name"] or ""
        if parent and parent == name:
            parent = ""
        if parent == "总公司" and "总公司" not in all_names_raw:
            parent = ""
        pair = (name, parent)
        if pair in seen_pairs:
            continue
        seen_pairs.add(pair)
        deduped.append({"name": name, "parent_name": parent or None})

    if not deduped:
        raise HTTPException(status_code=400, detail="Excel 在去重后无有效数据。")

    root_count = 0
    root_examples: list[str] = []
    for row in deduped:
        if not (row["parent_name"] or ""):
            root_count += 1
            if len(root_examples) < 8:
                root_examples.append(row["name"] or "")
    if root_count != 1:
        if root_count == 0:
            detail = "导入文件未识别到总公司：请至少保留1条上级公司为空的记录。"
        else:
            detail = (
                f"导入文件识别到 {root_count} 条总公司候选（上级公司为空），"
                f"仅允许1条。候选示例：{', '.join(root_examples)}"
            )
        raise HTTPException(status_code=400, detail=detail)
    return deduped


async def _clear_org_before_import(db: AsyncSession) -> None:
    logger.warning("【数据清空】组织架构导入：DELETE vehicle_device / vehicle / fleet / org_company，并清空用户 org_id。")
    await db.execute(text("DELETE FROM vehicle_device"))
    await db.execute(text("DELETE FROM vehicle"))
    await db.execute(text("DELETE FROM fleet"))
    await db.execute(text("UPDATE sys_user SET org_id = NULL"))
    await db.execute(text("DELETE FROM org_company"))
    if db.bind is not None and db.bind.dialect.name == "sqlite":
        exists = await db.execute(
            text("SELECT 1 FROM sqlite_master WHERE type='table' AND name='sqlite_sequence' LIMIT 1")
        )
        if exists.first():
            await db.execute(text("DELETE FROM sqlite_sequence WHERE name = 'org_company'"))


async def _import_rows_in_levels(db: AsyncSession, rows: list[dict[str, str | None]]) -> int:
    remaining = list(rows)
    name_to_ids: dict[str, list[int]] = {}
    imported = 0
    while remaining:
        progressed = False
        next_remaining: list[dict[str, str | None]] = []
        for row in remaining:
            name = row["name"] or ""
            parent_name = row["parent_name"]
            parent_ids = name_to_ids.get(parent_name or "", [])
            if parent_name and not parent_ids:
                next_remaining.append(row)
                continue
            obj = OrgCompany(
                name=name,
                legal_person=None,
                parent_id=(parent_ids[0] if parent_ids else None),
                contact_phone=None,
                address=None,
            )
            db.add(obj)
            await db.flush()
            obj.org_code = _gen_org_code(obj.id)
            await db.flush()
            name_to_ids.setdefault(name, []).append(obj.id)
            imported += 1
            progressed = True
        if not progressed:
            unresolved_msgs: list[str] = []
            imported_names = set(name_to_ids.keys())
            for r in next_remaining[:8]:
                cname = r["name"] or ""
                pname = r["parent_name"] or ""
                if pname and pname not in imported_names:
                    unresolved_msgs.append(f"{cname}(上级:{pname})")
                else:
                    unresolved_msgs.append(f"{cname}(层级循环或重复依赖)")
            msg = "；".join(unresolved_msgs) if unresolved_msgs else "未知异常"
            raise HTTPException(status_code=400, detail=f"已按“总公司->下级”逐层导入，但以下记录仍无法挂载：{msg}")
        remaining = next_remaining
    return imported


@router.get("/import-template")
async def download_company_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "公司信息导入"
    ws["A1"] = "公司名称"
    ws["B1"] = "上级公司"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    cd = "attachment; filename*=UTF-8''%E5%85%AC%E5%8F%B8%E4%BF%A1%E6%81%AF%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


@router.post("/import-excel")
async def import_company_excel(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式文件。")
    payload = await file.read()
    rows = _parse_import_rows(payload)
    rows = _dedupe_and_validate_import_rows(rows)
    await _clear_org_before_import(db)
    imported = await _import_rows_in_levels(db, rows)
    return {"ok": True, "imported": imported}


async def _load_all_map(db: AsyncSession) -> dict[int, OrgCompany]:
    r = await db.execute(select(OrgCompany))
    return {x.id: x for x in r.scalars().all()}


async def _descendant_ids(db: AsyncSession, root_id: int) -> set[int]:
    by_parent: dict[int | None, list[int]] = {}
    r = await db.execute(select(OrgCompany.id, OrgCompany.parent_id))
    for oid, pid in r.all():
        by_parent.setdefault(pid, []).append(oid)
    out: set[int] = set()
    stack = [root_id]
    while stack:
        cur = stack.pop()
        if cur in out:
            continue
        out.add(cur)
        for c in by_parent.get(cur, []):
            stack.append(c)
    return out


def _tree_nodes_for_ui(
    rows: list[OrgCompany],
    by_parent: dict[int | None, list[OrgCompany]],
    *,
    label_plain: bool = False,
) -> list[dict[str, Any]]:
    def build(node: OrgCompany) -> dict[str, Any]:
        kids = sorted(by_parent.get(node.id, []), key=lambda x: x.id)
        code = (node.org_code or "").strip()
        if label_plain:
            label = (node.name or "").strip()
        else:
            label = f"{node.name}（{code}）" if code else (node.name or "")
        return {
            "id": str(node.id),
            "label": label,
            "icon": "building",
            "children": [build(c) for c in kids],
        }

    roots = sorted(by_parent.get(None, []), key=lambda x: x.id)
    return [build(x) for x in roots]


@router.get("/tree")
async def org_tree(
    exclude_id: int | None = Query(None),
    label_plain: bool = Query(False),
    scope_org_tree: bool = Query(False),
    x_org_id: str | None = Header(None, alias="X-Org-Id"),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(OrgCompany).order_by(OrgCompany.id))
    rows = list(r.scalars().all())

    do_scope = wants_org_tree_scope(scope_org_tree, x_org_id)
    if do_scope:
        root = require_x_org_id_header(x_org_id)
        subtree = await collect_org_company_subtree_ids(db, root)
        rows = [x for x in rows if x.id in subtree]

    if exclude_id is not None:
        bad = await _descendant_ids(db, exclude_id)
        pruned = [x for x in rows if x.id not in bad]
        valid_ids = {x.id for x in pruned}
        by_parent: dict[int | None, list[OrgCompany]] = {}
        for row in pruned:
            pid = row.parent_id if row.parent_id in valid_ids else None
            by_parent.setdefault(pid, []).append(row)
        return {"tree": _tree_nodes_for_ui(pruned, by_parent, label_plain=label_plain)}

    valid_ids = {x.id for x in rows}
    by_parent = {}
    for row in rows:
        pid = row.parent_id if (not do_scope or row.parent_id in valid_ids) else None
        by_parent.setdefault(pid, []).append(row)
    return {"tree": _tree_nodes_for_ui(rows, by_parent, label_plain=label_plain)}


@router.get("/parent-options")
async def parent_options(exclude_id: int | None = Query(None), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(OrgCompany).order_by(OrgCompany.id))
    all_rows = list(r.scalars().all())
    bad: set[int] = set()
    if exclude_id is not None:
        bad = await _descendant_ids(db, exclude_id)

    roots = [x for x in all_rows if x.parent_id is None]
    has_root = len(roots) >= 1
    root_id = roots[0].id if roots else None
    pmap = {x.id: x.parent_id for x in all_rows}

    def depth(i: int) -> int:
        d = 0
        p = pmap.get(i)
        while p is not None and p in pmap:
            d += 1
            p = pmap.get(p)
        return d

    options: list[dict[str, Any]] = []
    can_show_head_option = (not has_root) or (
        exclude_id is not None and root_id is not None and exclude_id == root_id
    )
    if can_show_head_option:
        options.append({"id": None, "name": "无（总公司）", "depth": 0})

    for row in all_rows:
        if row.id in bad:
            continue
        options.append(
            {"id": row.id, "name": ("　" * depth(row.id)) + row.name, "depth": depth(row.id)}
        )
    return {"options": options}


@router.get("/companies", response_model=dict)
async def company_list(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    org_code: str | None = None,
    name: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    conds = []
    if org_code:
        conds.append(OrgCompany.org_code.ilike(f"%{org_code.strip()}%"))
    if name:
        conds.append(OrgCompany.name.ilike(f"%{name.strip()}%"))
    count_stmt = select(func.count()).select_from(OrgCompany)
    if conds:
        count_stmt = count_stmt.where(*conds)
    total = (await db.execute(count_stmt)).scalar() or 0
    q = select(OrgCompany).order_by(OrgCompany.id.desc()).offset((page - 1) * page_size).limit(page_size)
    if conds:
        q = q.where(*conds)
    rows = (await db.execute(q)).scalars().all()
    pmap = await _load_all_map(db)

    items = []
    for x in rows:
        if x.parent_id is None:
            pn = "总公司"
        elif x.parent_id in pmap:
            pn = pmap[x.parent_id].name
        else:
            pn = "—"
        items.append(
            {
                "id": x.id,
                "org_code": x.org_code,
                "name": x.name,
                "legal_person": x.legal_person,
                "parent_id": x.parent_id,
                "parent_name": pn,
                "contact_phone": x.contact_phone,
                "address": x.address,
                "created_at": x.created_at.isoformat() if x.created_at else None,
            }
        )
    return {"total": total, "items": items, "page": page, "page_size": page_size}


@router.get("/companies/{company_id}")
async def company_detail(company_id: int, db: AsyncSession = Depends(get_db)):
    x = (await db.execute(select(OrgCompany).where(OrgCompany.id == company_id))).scalar_one_or_none()
    if not x:
        raise HTTPException(404, "组织不存在")
    pmap = await _load_all_map(db)
    pn = pmap[x.parent_id].name if x.parent_id and x.parent_id in pmap else None
    return {
        "id": x.id,
        "org_code": x.org_code,
        "name": x.name,
        "legal_person": x.legal_person,
        "parent_id": x.parent_id,
        "parent_name": pn,
        "contact_phone": x.contact_phone,
        "address": x.address,
        "created_at": x.created_at.isoformat() if x.created_at else None,
    }


async def _other_root_count(db: AsyncSession, exclude_id: int | None) -> int:
    stmt = select(func.count()).select_from(OrgCompany).where(OrgCompany.parent_id.is_(None))
    if exclude_id is not None:
        stmt = stmt.where(OrgCompany.id != exclude_id)
    return (await db.scalar(stmt)) or 0


async def _validate_single_head_company(db: AsyncSession, want_parent_id: int | None, exclude_id: int | None) -> None:
    if want_parent_id is not None:
        return
    if await _other_root_count(db, exclude_id) >= 1:
        raise HTTPException(
            status_code=400,
            detail="系统中仅能有一个总公司（无上级单位）。请从「上级单位」中选择已有公司，或调整原总公司后再设。",
        )


async def _validate_parent(db: AsyncSession, parent_id: int | None, exclude_id: int | None) -> None:
    if parent_id is None:
        return
    pr = await db.execute(select(OrgCompany).where(OrgCompany.id == parent_id))
    if pr.scalar_one_or_none() is None:
        raise HTTPException(400, "上级单位不存在")
    if exclude_id is not None:
        bad = await _descendant_ids(db, exclude_id)
        if parent_id in bad or parent_id == exclude_id:
            raise HTTPException(400, "不能将上级设为自己或自己的下级")


async def _resolve_jt_fid(db: AsyncSession, parent_id: int | None) -> tuple[int | None, str | None]:
    if parent_id is None:
        return 0, None
    pr = await db.execute(select(OrgCompany).where(OrgCompany.id == parent_id))
    parent = pr.scalar_one_or_none()
    if parent is None or parent.jt808_group_id is None:
        return None, "上级公司尚未同步到 JT808（缺 jt808_group_id），已跳过分组联动"
    return int(parent.jt808_group_id), None


@router.post("/companies")
async def company_create(
    body: OrgCompanyCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    await _validate_single_head_company(db, body.parent_id, None)
    await _validate_parent(db, body.parent_id, None)
    row = OrgCompany(
        name=body.name.strip(),
        legal_person=(body.legal_person or "").strip() or None,
        parent_id=body.parent_id,
        contact_phone=(body.contact_phone or "").strip() or None,
        address=(body.address or "").strip() or None,
    )
    db.add(row)
    await db.flush()
    row.org_code = _gen_org_code(row.id)
    fid, jt_warn = await _resolve_jt_fid(db, body.parent_id)
    cid, cname, ccode = row.id, row.name, row.org_code
    await db.commit()

    queued = False
    if fid is not None:
        background_tasks.add_task(jt808_group.bg_create, cid, cname, fid)
        queued = True
    return {"id": cid, "org_code": ccode,
            "jt808_sync": ("queued" if queued else "skipped"), "jt808_warn": jt_warn}


@router.put("/companies/{company_id}")
async def company_update(
    company_id: int,
    body: OrgCompanyUpdate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(OrgCompany).where(OrgCompany.id == company_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "组织不存在")
    patch = body.model_dump(exclude_unset=True)
    old_name = row.name
    name_changed = False
    parent_changed = False
    if "parent_id" in patch:
        await _validate_single_head_company(db, patch["parent_id"], company_id)
        await _validate_parent(db, patch["parent_id"], company_id)
        parent_changed = patch["parent_id"] != row.parent_id
        row.parent_id = patch["parent_id"]
    if "name" in patch and patch["name"] is not None:
        new_name = patch["name"].strip()
        name_changed = new_name != old_name
        row.name = new_name
    if "legal_person" in patch:
        row.legal_person = (patch["legal_person"] or "").strip() or None
    if "contact_phone" in patch:
        row.contact_phone = (patch["contact_phone"] or "").strip() or None
    if "address" in patch:
        row.address = (patch["address"] or "").strip() or None
    row.updated_at = datetime.now(timezone.utc)

    do_edit = bool(row.jt808_group_id and (name_changed or parent_changed))
    new_fid = None
    if do_edit:
        new_fid, _ = await _resolve_jt_fid(db, row.parent_id)
    gid, new_name = row.jt808_group_id, row.name
    await db.commit()

    if do_edit:
        background_tasks.add_task(jt808_group.bg_edit, gid, new_name, new_fid)
    return {"ok": True, "jt808_sync": ("queued" if do_edit else None)}


@router.delete("/companies/{company_id}")
async def company_delete(
    company_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    row = (await db.execute(select(OrgCompany).where(OrgCompany.id == company_id))).scalar_one_or_none()
    if not row:
        raise HTTPException(404, "组织不存在")
    ch = await db.scalar(select(func.count()).select_from(OrgCompany).where(OrgCompany.parent_id == company_id))
    if ch and ch > 0:
        raise HTTPException(400, "存在下级组织，无法删除")
    fv = await db.scalar(select(func.count()).select_from(Fleet).where(Fleet.company_id == company_id))
    if fv and fv > 0:
        raise HTTPException(400, "该组织下存在车队，无法删除")
    vv = await db.scalar(select(func.count()).select_from(Vehicle).where(Vehicle.company_id == company_id))
    if vv and vv > 0:
        raise HTTPException(400, "该组织下存在车辆，无法删除")

    gid = row.jt808_group_id
    await db.execute(delete(OrgCompany).where(OrgCompany.id == company_id))
    await db.commit()
    if gid:
        background_tasks.add_task(jt808_group.bg_delete, gid)
    return {"ok": True, "jt808_sync": ("queued" if gid else None)}
